import sys, time, os, re
#from bs4 import BeautifulSoup
#from bs4 import UnicodeDammit
'''for excel'''
from openpyxl import Workbook
from htmllib5 import etree_lxml
import urllib2

start = time.clock()
wb = Workbook()
ws = wb.get_active_sheet()

num = '0'

address = ('http://www.cantonese.sheik.co.uk/scripts/wordlist.htm?action=&wordtype=0&level=-1&page='+num)
html = urllib2.urlopen(address).read() 
source = etree.Element(html)

print source


'''
# special method to extract td valign="top" then reiterate for 3, 9, 15, etc. to get definitions
def only_valign(tag):
	return tag.has_key('valign') and not tag.has_key('class')
	
status = "Scraping page " + str(increment) + "..."
#print '\r' + status

	
word = soup.find_all("td",{"class":"wl_uni"})
jyutping = soup.find_all("span",{"listjyutping"}) 
pinyin = soup.find_all("span",{"listpinyin"}) 
defn = soup.find_all(only_valign)

for counter in range(0,20):
	num = str((increment)*20+(counter+1))
	defn_counter = 3+(counter*6)
	
	cell1 = ws.cell('A'+num) 
	cell2 = ws.cell('B'+num) 
	cell3 = ws.cell('C'+num)
	cell4 = ws.cell('D'+num)
	try:
		#print word[counter].get_text()
		cell1.value = word[counter].get_text()
	except UnicodeEncodeError:
		print "-----------NA-----------"
		cell1.value = "check "+num
	
	try:
		#print jyutping[counter].get_text()
		cell2.value = jyutping[counter].get_text()
	except UnicodeEncodeError:
		cell2.value = "check "+num
	
	try:
		#print pinyin[counter].get_text()
		cell3.value = pinyin[counter].get_text()
	except UnicodeEncodeError:
		cell3.value =  "check "+num
	
	
	try:
		#print defn[defn_counter].get_text()
		cell4.value = defn[defn_counter].get_text()
	except UnicodeEncodeError:
		cell4.value =  "check "+num

wb.save('test.xlsx')

end = time.clock()
#print end-start
'''