import urllib2, sys, time
from bs4 import BeautifulSoup
from bs4 import UnicodeDammit
from openpyxl import Workbook # for exporting to excel

start = time.clock() # for time tracking
wb = Workbook()
ws = wb.get_active_sheet()

def parse():
	for counter in range(0,20): # 20 words per page
	
		# adjust for what is actually in the html source, used trial and error
		num = str((increment)*20+(counter+1))
		defn_counter = 3+(counter*6)
		
		# set variables for excel cell values to write to		
		cell1 = ws.cell('A'+num) 
		cell2 = ws.cell('B'+num) 
		cell3 = ws.cell('C'+num)
		cell4 = ws.cell('D'+num)
		
		# parse the word
		try:
			#print word[counter].get_text()
			cell1.value = word[counter].get_text()
		except IndexError:
			print "End of file reached."
			break
		except UnicodeEncodeError:
			print "-----------NA-----------"
			cell1.value = "check "+num
		
		# parse the jyutping pronounciation
		try:
			#print jyutping[counter].get_text()
			cell2.value = jyutping[counter].get_text()
		except UnicodeEncodeError:
			cell2.value = "check "+num
		
		# parse the pinyin
		try:
			#print pinyin[counter].get_text()
			cell3.value = pinyin[counter].get_text()
		except UnicodeEncodeError:
			cell3.value =  "check "+num
		
		# parse the definition
		try:
			#print defn[defn_counter].get_text()
			cell4.value = defn[defn_counter].get_text()
		except UnicodeEncodeError:
			cell4.value =  "check "+num
			
for increment in range(0,3000): #there's currently around 2882 pages that are updated daily, but if no more pages beyond that exist the program will exit
	
	# grab the source
	num = str(increment)
	address = ('http://www.cantonese.sheik.co.uk/scripts/wordlist.htm?action=&wordtype=0&level=-1&page='+num)
	html = urllib2.urlopen(address).read() 
	soup = BeautifulSoup(html) 

	# special method to extract td valign="top" then reiterate for 3, 9, 15, etc. to get definitions
	def only_valign(tag):
		return tag.has_key('valign') and not tag.has_key('class')
	
	# show what page we're parsing/scraping
	print "Working on page",increment
	
	word = soup.find_all("td",{"class":"wl_uni"})
	jyutping = soup.find_all("span",{"listjyutping"}) 
	pinyin = soup.find_all("span",{"listpinyin"}) 
	defn = soup.find_all(only_valign)
	
	parse()

	wb.save('cantodict.xlsx')
	
# write all values to excel file
wb.save('cantodict.xlsx')

# print out total time required to parse
end = time.clock()
print "\nTotal time elapsed: ", end-start, " ms."
print "Average time per page = ", (end-start)/(increment+1), " ms."